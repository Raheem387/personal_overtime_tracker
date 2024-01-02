import openpyxl
from datetime import datetime

# Get user input
location = input("Enter location: ")
date_str = input("Enter date (YYYY-MM-DD): ")
hours_worked = input("Enter hours worked: ")
report_number = input("Enter report number: ")

# Parse date input
try:
    date = datetime.strptime(date_str, "%Y-%m-%d").date()
except ValueError:
    print("Invalid date format. Please use YYYY-MM-DD.")
    exit()

# Create or load the Excel spreadsheet
excel_file_path = "reports.xlsx"
try:
    workbook = openpyxl.load_workbook(excel_file_path)
except FileNotFoundError:
    workbook = openpyxl.Workbook()

# Select the active sheet
sheet = workbook.active

# Add headers if the sheet is empty
if sheet.max_row == 1:
    sheet['A1'] = 'Location'
    sheet['B1'] = 'Date'
    sheet['C1'] = 'Hours Worked'
    sheet['D1'] = 'Report Number'

# Append user input to the spreadsheet
row = (location, date, hours_worked, report_number)
sheet.append(row)

# Save the workbook
workbook.save(excel_file_path)

print(f"Data added to {excel_file_path}")
